from openpyxl import load_workbook

FILE = "classement.xlsx"

def load_or_create():
    try:
        wb = load_workbook(FILE)
        ws = wb.active
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "elo"])
        wb.save(FILE)
    return wb, ws

def get_players(ws):
    players = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            players[row[0]] = row[1]
    return players

def save(ws, players):
    ws.delete_rows(2, ws.max_row)
    for name, elo in sorted(players.items(), key=lambda x: x[1], reverse=True):
        ws.append([name, elo])

def main():
    wb, ws = load_or_create()
    players = get_players(ws)

    name = input("Nom du joueur : ").strip()

    if name not in players:
        print(f"❌ Le joueur '{name}' n'existe pas dans le classement.")
        return

    new_elo_input = input("Nouvel Elo (laisser vide pour supprimer) : ").strip()

    if not new_elo_input:
        # Demander confirmation avant suppression
        confirm = input(f"⚠️ Êtes-vous sûr de vouloir supprimer '{name}' ? (Oui/Non) : ").strip().lower()
        if confirm == "oui":
            del players[name]
            print(f"✅ Le joueur '{name}' a été supprimé.")
        else:
            print("❌ Suppression annulée.")
            return
    else:
        try:
            new_elo = int(new_elo_input)
            players[name] = new_elo
            print(f"✅ Elo de '{name}' mis à jour à {new_elo}.")
        except ValueError:
            print("❌ Veuillez entrer un nombre valide pour l'Elo.")
            return

    save(ws, players)
    wb.save(FILE)

    print("\n🏆 Classement mis à jour :\n")
    for n, e in sorted(players.items(), key=lambda x: x[1], reverse=True):
        print(f"{n}: {e}")

if __name__ == "__main__":
    main()