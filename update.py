import os
import math
from openpyxl import Workbook, load_workbook

K = 30
FILE = "classement.xlsx"


# -------------------------
# Initialisation Excel
# -------------------------
def init_file():
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Classement"

        ws.append(["name", "elo"])

        wb.save(FILE)


# -------------------------
# Chargement joueurs
# -------------------------
def load_players():
    wb = load_workbook(FILE)
    ws = wb.active

    players = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, elo = row
        players[name] = elo

    return wb, ws, players


# -------------------------
# Sauvegarde
# -------------------------
def save_players(wb, ws, players):
    ws.delete_rows(2, ws.max_row)

    # tri par elo
    sorted_players = sorted(players.items(), key=lambda x: x[1], reverse=True)

    for name, elo in sorted_players:
        ws.append([name, round(elo)])

    wb.save(FILE)


# -------------------------
# Elo
# -------------------------
def expected_score(rA, rB):
    return 1 / (1 + 10 ** ((rB - rA) / 400))

def expected_score_3(rA, rB, rC):
    return 1 / (1 + 10 ** (((rB+rC)/2 - rA) / 400))


def avg_elo(players, names):
    return sum(players[n] for n in names) / len(names)


def ensure_players_exist(players, names):
    for name in names:
        if name not in players:
            print(f"➕ Nouveau joueur détecté: {name} (Elo = 1000)")
            players[name] = 1000


def update_elo(players, winners, losers, neutrals):
    if len(neutrals) > 0:
        avg_win = avg_elo(players, winners)
        avg_neutral = avg_elo(players, neutrals)
        avg_lose = avg_elo(players, losers)
        
        """
        exp_WN = expected_score(avg_win, avg_neutral)
        exp_WL = expected_score(avg_win, avg_lose)

        exp_NW = 1- exp_WN
        exp_NL = expected_score(avg_neutral, avg_lose)

        exp_LW = 1 - exp_WL
        exp_LN = 1 - exp_NL
        """

        exp_win = expected_score_3(avg_win, avg_neutral, avg_lose)
        exp_neutral = expected_score_3(avg_neutral, avg_neutral, avg_win)
        exp_lose = expected_score_3(avg_lose, avg_win, avg_lose)

        for p in winners:
            players[p] += K * (1 - exp_win)
        
        for p in neutrals:
            players[p] += K * (0.5 - exp_neutral)

        for p in losers:
            players[p] += K * (0 - exp_lose)
        
    else : 
        avg_win = avg_elo(players, winners)
        avg_lose = avg_elo(players, losers)
        

        exp_win = expected_score(avg_win, avg_lose)
        exp_lose = expected_score(avg_lose, avg_win)

        for p in winners:
            players[p] += K * (1 - exp_win)

        for p in losers:
            players[p] += K * (0 - exp_lose)

    


# -------------------------
# Input utilisateur
# -------------------------
def ask_players():
    winners = input("Gagnants (séparés par espace) : ").split()
    losers = input("Perdants (séparés par espace) : ").split()
    neutrals = input("Neutres (optionnel) : ").split()

    return winners, losers, neutrals


# -------------------------
# Main
# -------------------------
def main():
    init_file()

    wb, ws, players = load_players()

    winners, losers, neutrals = ask_players()

    all_players = winners + losers + neutrals

    ensure_players_exist(players, all_players)

    update_elo(players, winners, losers, neutrals)

    save_players(wb, ws, players)

    print("\n🏆 Classement mis à jour :\n")
    for name, elo in sorted(players.items(), key=lambda x: x[1], reverse=True):
        print(f"{name}: {round(elo)}")


if __name__ == "__main__":
    main()